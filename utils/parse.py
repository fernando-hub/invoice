
import configparser
from doctest import OutputChecker
from openpyxl import Workbook
from utils.Invoice import *
import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import logging
from utils.Email import *

rows=[]

#now we will Create and configure logger 
logging.basicConfig(filename="std.log", 
					format='%(asctime)s %(message)s', 
					filemode='w') 

#Let us Create an object 
logger=logging.getLogger() 

#Now we are going to Set the threshold of logger to DEBUG 
logger.setLevel(logging.DEBUG) 

def parsefile():

    read_config = configparser.ConfigParser()
    read_config.read("app.config")
    nameFile = read_config.get("DEFAULT", "locationfile")
    
    
    file1 = open(nameFile,"r+")   
   

    # Using readlines()    
    Lines = file1.readlines()
 
    count = 0
    jump= 0
    saveLine=False
    # Strips the newline character
    for line in Lines:
        count += 1
        #print("Line{}: {}".format(count, line.strip()))
        if  jump>0:
            jump-=1
            continue

        if  jump==0 and saveLine:
            saveLine=False
            invoicehead= splitvalues(line.strip())
            #encabezado=encabezado.split('  ')
            #print (encabezado)
            #invoicehead=line.strip().split('      ')
            searchDetail = True
            invo=Invoice()
            countDetail=1

            while searchDetail:                
                #searchDetail != firstword(Lines[count+1+countDetail].strip())=="Total:"
                if  firstword(Lines[count+countDetail].strip())=="---------------" or firstword(Lines[count+countDetail].strip())=="QSL" or firstword(Lines[count+countDetail].strip())=="":
                    searchDetail=False                    
                else:
                    
                    row1=Lines[count+countDetail].strip().split()
                    #print (row1) 
                    try:
                        invo.CustomerName=invoicehead[0]        
                    except IndexError:
                        pass
                    if len(invoicehead)>1:
                        invo.CustomerNumber=invoicehead[1]
                        try:
                            invo.City=invoicehead[2]
                        except IndexError:                        
                            pass
                        try:
                            invo.State=invoicehead[3]
                        except IndexError:                        
                            pass
                        try:                            
                            invo.Username=invoicehead[4]
                        except IndexError:                        
                            pass
                    invo.Number=Lines[count+countDetail].strip().split()[0]   
                    try:
                        invo.Type=row1[1]+' '+row1[2] 
                    except:
                        pass
                    try:                            
                        invo.DueDate= validdate(Lines[count+countDetail].strip().split()[3])
                        invo.CurrentDue=Lines[count+countDetail].strip().split()[4]
                        invo.Outstanding=Lines[count+countDetail].strip().split()[5]
                        invo.Due1=Lines[count+countDetail].strip().split()[6]
                        invo.Due6=Lines[count+countDetail].strip().split()[6]
                        invo.Due16=Lines[count+countDetail].strip().split()[6]
                        invo.Due31=Lines[count+countDetail].strip().split()[6]
                        invo.Due61=Lines[count+countDetail].strip().split()[6]
                        invo.Due90=Lines[count+countDetail].strip().split()[6]
                        
                                                
                    except IndexError:
                        #print("no valido")
                        pass
                    #print(invoicehead)
                    #invo.ShowInvoice()
                    if(invo.CustomerName!="" and invo.CustomerName!="Customer Type Total:"):
                        rows.append(invo)

                #print(Lines[count+countDetail].strip())
                #firstword(Lines[count+1+countDetail].strip())
                countDetail+=1

        words=line.strip().split()
        if len(words)>0 and words[0]=="Invoice":            
            #words2=Lines[count].strip().split()
            #if len(words2)>0 and words2[0]=="Number":
            if firstword(Lines[count].strip())=="Number":
                jump=2
                saveLine=True
                continue
            #else:
                #print(words2)
        if len(words)>0 and words[0]=="Customer" and words[1]=="Type" and words[2]=="Total:":
            jump=1
            saveLine=True
            continue
    return rows                       

def createExcel(rows):
        
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws.cell(row=1, column=1).value = "Customer Name"	
        ws.cell(row=1, column=2).value = "Customer Number"	
        ws.cell(row=1, column=3).value = "Invoice Number"	
        ws.cell(row=1, column=4).value = "Type"	
        ws.cell(row=1, column=5).value = "Due Date"
        ws.cell(row=1, column=6).value = "Current Due"	
        ws.cell(row=1, column=7).value = "Due 1 to 5"	
        ws.cell(row=1, column=8).value = "Due 6 to 15"	
        ws.cell(row=1, column=9).value = "Due 16 to 30"	
        ws.cell(row=1, column=10).value = "Due 31 to 60"
        ws.cell(row=1, column=11).value = "Due 61 to 90"	
        ws.cell(row=1, column=12).value = "Due 90+"	
        ws.cell(row=1, column=13).value = "Outstanding Amount"        
        ws.cell(row=1, column=14).value = "City"
        ws.cell(row=1, column=15).value = "State"	
        ws.cell(row=1, column=16).value = "Username"
        
        colorFill = PatternFill(start_color='C0C0C0',
                   end_color='C0C0C0',
                   fill_type='solid')
        
        
        
        ws.cell(row=1, column=1).fill=colorFill
        ws.cell(row=1, column=2).fill=colorFill
        ws.cell(row=1, column=3).fill=colorFill
        ws.cell(row=1, column=4).fill=colorFill
        ws.cell(row=1, column=5).fill=colorFill
        ws.cell(row=1, column=6).fill=colorFill
        ws.cell(row=1, column=7).fill=colorFill
        ws.cell(row=1, column=8).fill=colorFill
        ws.cell(row=1, column=9).fill=colorFill
        ws.cell(row=1, column=10).fill=colorFill
        ws.cell(row=1, column=11).fill=colorFill
        ws.cell(row=1, column=12).fill=colorFill
        ws.cell(row=1, column=13).fill=colorFill    
        ws.cell(row=1, column=14).fill=colorFill
        ws.cell(row=1, column=15).fill=colorFill
        ws.cell(row=1, column=16).fill=colorFill
        

        rowcount=2
        for item in rows:
            ws.cell(row=rowcount,column=1).value=item.CustomerName
            ws.cell(row=rowcount,column=2).value=item.CustomerNumber
            ws.cell(row=rowcount,column=3).value=item.Number
            ws.cell(row=rowcount,column=4).value=item.Type
            ws.cell(row=rowcount,column=5).value=item.DueDate
            ws.cell(row=rowcount,column=6).value=Curren(item.CurrentDue)
            ws.cell(row=rowcount,column=7).value=Curren(item.Due1)
            ws.cell(row=rowcount,column=8).value=Curren(item.Due6)
            ws.cell(row=rowcount,column=9).value=Curren(item.Due16)
            ws.cell(row=rowcount,column=10).value=Curren(item.Due31)
            ws.cell(row=rowcount,column=11).value=Curren(item.Due61)
            ws.cell(row=rowcount,column=12).value=Curren(item.Due90)
            ws.cell(row=rowcount,column=13).value=Curren(item.Outstanding)
            ws.cell(row=rowcount,column=14).value=item.City
            ws.cell(row=rowcount,column=15).value=item.State
            ws.cell(row=rowcount,column=16).value=item.Username
            rowcount+=1

        # Save the file
        read_config = configparser.ConfigParser()
        read_config.read("app.config")
        outputFile = read_config.get("DEFAULT", "locationexport")
        mailTo=read_config.get("DEFAULT", "Email")
        try:
            wb.save(outputFile+"Output.xlsx")
            print ("File successful created, "+outputFile+"Output.xlsx")
            # send email
            #try:
            mailing=MAIL(mailTo,outputFile+"Output.xlsx")
            mailing.send()
            #except :
                #print ("Something went wrong with the mail "+mailTo)
                #logger.debug("Something went wrong with the mail "+mailTo) 
        except FileNotFoundError:
            print ("No such file or directory, please update to valid path")
            logger.debug("No such file or directory, please update to valid path "+outputFile) 
            pass

def firstword(line):
    words=line.split()
    if len(words)>0:
        return words[0]
    else:
        return ""

def validdate(datestring):
    date=datestring.split('-')
    year=date[2]
    if int(year) < 8:
        year=int(year)+2000
    else:
        year=int(year)+1900

    return printMonth(date[1])+'/'+date[0]+'/'+str(year)

def printMonth(num):
    if num == "JAN":
        month ='01'
    elif num == "FEB":
        month = '02'
    elif num == "MAR":
        month = '03'
    elif num == "APR":
        month= '04'
    elif num == "MAY":
        month= '05'
    elif num == "JUN":
        month = '06'
    elif num == "JUL":
        month = '07'
    elif num == "AUG":
        month = '08'
    elif num == "SEP":
        month= '09'
    elif num == "OCT":
        month= '10'
    elif num == "NOV":
        month= '11'
    elif num == "DEC":
        month= '12'
    return month


def splitvalues(stringin):
    stringin=stringin.split('  ')
    response=[]

    for x in stringin:
        if x.strip() != "":
            response.append(x)
    return response
       
def Curren(mount):
    if mount=="":
        mount=0.00
        
    return "$"+str(mount)        
